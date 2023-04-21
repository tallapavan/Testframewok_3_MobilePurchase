import inspect
import logging
import csv
from openpyxl import load_workbook


class Utils():

    def custom_logger():
        # Step-1
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.DEBUG)
        # Step-2
        file_handler = logging.FileHandler('automation.log', mode= 'a')
        # Step-3
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s: %(message)s',
                                      datefmt='%Y-%m-%d %H:%M:%S')
        # Step-4
        file_handler.setFormatter((formatter))
        # Step-5
        logger.addHandler(file_handler)
        return logger

    def read_excel_data(file_path, Sheet_name):
        excel_data = []
        workbook = load_workbook(file_path)
        sheet = workbook[Sheet_name]

        max_row = sheet.max_row
        max_col = sheet.max_column
        for r in range(2, max_row + 1):
            row_data = []
            for c in range(1, max_col + 1):
                row_data.append(sheet.cell(row=r, column=c).value)
            excel_data.append(row_data)
        return excel_data

    def read_csv_data(file_path):
        csv_list = []
        # open csv file
        csv_data = open(file_path, 'r')
        # read csv
        reader = csv.reader(csv_data)
        next(reader)    # this step will skip the hearder data and read next line

        for row in reader:
            csv_list.append(row)
        return csv_list
